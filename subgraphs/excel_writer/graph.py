"""
Excel Writer Subgraph Builder

Saves job application data to Excel tracking spreadsheet.

Graph Flow:
```
    START
      │
      ▼
┌─────────────────────────┐
│    prepare_record       │  ← Create ApplicationRecord from inputs
└───────────┬─────────────┘
            │
            ▼
┌─────────────────────────┐
│ load_or_create_workbook │  ← Load existing or create new Excel
└───────────┬─────────────┘
            │
            ▼
┌─────────────────────────┐
│     write_record        │  ← Write application to spreadsheet
└───────────┬─────────────┘
            │
            ▼
┌─────────────────────────┐
│   format_spreadsheet    │  ← Apply formatting/styling
└───────────┬─────────────┘
            │
            ▼
┌─────────────────────────┐
│     save_workbook       │  ← Save and close file
└───────────┬─────────────┘
            │
            ▼
          END
```
"""

from typing import Dict, Any, Optional
from pathlib import Path
from langgraph.graph import StateGraph, START, END

from subgraphs.excel_writer.state import (
    ExcelWriterState,
    ApplicationStatus,
    ApplicationSource,
    ApplicationRecord,
    SpreadsheetConfig,
    SPREADSHEET_COLUMNS,
    generate_application_id,
    get_today
)
from subgraphs.excel_writer.nodes import (
    prepare_record,
    load_or_create_workbook,
    write_record,
    format_spreadsheet,
    save_workbook
)


def build_excel_writer_graph() -> StateGraph:
    """
    Build the Excel Writer subgraph.
    
    Returns:
        Compiled StateGraph for Excel writing
    """
    graph = StateGraph(ExcelWriterState)
    
    # Add nodes
    graph.add_node("prepare_record", prepare_record)
    graph.add_node("load_or_create_workbook", load_or_create_workbook)
    graph.add_node("write_record", write_record)
    graph.add_node("format_spreadsheet", format_spreadsheet)
    graph.add_node("save_workbook", save_workbook)
    
    # Add edges
    graph.add_edge(START, "prepare_record")
    graph.add_edge("prepare_record", "load_or_create_workbook")
    graph.add_edge("load_or_create_workbook", "write_record")
    graph.add_edge("write_record", "format_spreadsheet")
    graph.add_edge("format_spreadsheet", "save_workbook")
    graph.add_edge("save_workbook", END)
    
    return graph.compile()


def create_excel_writer_subgraph():
    """Create and return the compiled Excel Writer subgraph."""
    return build_excel_writer_graph()


# ============================================================================
# CONVENIENCE FUNCTIONS
# ============================================================================

def save_application_to_excel(
    structured_jd=None,
    resume_json=None,
    resume_score: float = 0.0,
    cover_letter_score: float = 0.0,
    application_record: ApplicationRecord = None,
    file_path: str = "job_applications.xlsx",
    status: ApplicationStatus = ApplicationStatus.NOT_APPLIED,
    source: ApplicationSource = ApplicationSource.COMPANY_WEBSITE,
    notes: str = "",
    why_interested: str = ""
) -> Dict[str, Any]:
    """
    Save a job application to Excel spreadsheet.
    
    Args:
        structured_jd: StructuredJD object with job details
        resume_json: ResumeJSON object (optional)
        resume_score: Score from compliance check (0-100)
        cover_letter_score: Score from compliance check (0-100)
        application_record: Pre-filled ApplicationRecord (optional)
        file_path: Path to Excel file
        status: Application status
        source: How the job was found
        notes: Additional notes
        why_interested: Reason for interest
    
    Returns:
        Dict with file_path, row_number, and write status
    """
    graph = create_excel_writer_subgraph()
    
    # Create or update record
    if not application_record:
        application_record = ApplicationRecord(
            status=status,
            source=source,
            notes=notes,
            why_interested=why_interested
        )
    
    config = SpreadsheetConfig(file_path=file_path)
    
    initial_state = {
        "structured_jd": structured_jd,
        "resume_json": resume_json,
        "resume_score": resume_score,
        "cover_letter_score": cover_letter_score,
        "application_record": application_record,
        "config": config
    }
    
    result = graph.invoke(initial_state)
    
    return {
        "file_path": result.get("file_path"),
        "row_number": result.get("row_number"),
        "write_complete": result.get("write_complete", False),
        "application_record": result.get("application_record"),
        "error": result.get("error_message")
    }


def quick_save_application(
    company: str,
    role: str,
    job_url: str = "",
    status: ApplicationStatus = ApplicationStatus.APPLIED,
    source: ApplicationSource = ApplicationSource.LINKEDIN,
    file_path: str = "job_applications.xlsx",
    notes: str = ""
) -> Dict[str, Any]:
    """
    Quick save an application with minimal info.
    
    Args:
        company: Company name
        role: Job title
        job_url: URL to job posting
        status: Application status
        source: How the job was found
        file_path: Path to Excel file
        notes: Additional notes
    
    Returns:
        Dict with file_path and status
    """
    record = ApplicationRecord(
        company=company,
        role=role,
        job_url=job_url,
        status=status,
        source=source,
        notes=notes,
        date_applied=get_today() if status != ApplicationStatus.NOT_APPLIED else ""
    )
    
    return save_application_to_excel(
        application_record=record,
        file_path=file_path
    )


def update_application_status(
    file_path: str,
    application_id: str,
    new_status: ApplicationStatus,
    notes: str = ""
) -> Dict[str, Any]:
    """
    Update an existing application's status.
    
    Args:
        file_path: Path to Excel file
        application_id: ID of application to update
        new_status: New status
        notes: Additional notes to append
    
    Returns:
        Dict with update status
    """
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Find the row with this application ID
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == application_id:
                # Update status (column 9)
                ws.cell(row=row_idx, column=9, value=new_status.value)
                # Update last updated (column 8)
                ws.cell(row=row_idx, column=8, value=get_today())
                # Append notes if provided
                if notes:
                    existing_notes = ws.cell(row=row_idx, column=26).value or ""
                    new_notes = f"{existing_notes}\n[{get_today()}] {notes}" if existing_notes else notes
                    ws.cell(row=row_idx, column=26, value=new_notes)
                
                wb.save(file_path)
                return {
                    "updated": True,
                    "row_number": row_idx,
                    "new_status": new_status.value
                }
        
        return {"updated": False, "error": f"Application {application_id} not found"}
        
    except Exception as e:
        return {"updated": False, "error": str(e)}


def get_application_stats(file_path: str) -> Dict[str, Any]:
    """
    Get statistics from the application tracker.
    
    Args:
        file_path: Path to Excel file
    
    Returns:
        Dict with statistics
    """
    try:
        from openpyxl import load_workbook
        from collections import Counter
        
        if not Path(file_path).exists():
            return {"error": "File not found", "total": 0}
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        total = ws.max_row - 1  # Exclude header
        if total <= 0:
            return {"total": 0, "by_status": {}, "by_source": {}}
        
        # Count by status
        statuses = []
        sources = []
        companies = []
        
        for row_idx in range(2, ws.max_row + 1):
            status = ws.cell(row=row_idx, column=9).value
            source = ws.cell(row=row_idx, column=10).value
            company = ws.cell(row=row_idx, column=2).value
            
            if status:
                statuses.append(status)
            if source:
                sources.append(source)
            if company:
                companies.append(company)
        
        return {
            "total": total,
            "by_status": dict(Counter(statuses)),
            "by_source": dict(Counter(sources)),
            "unique_companies": len(set(companies)),
            "top_companies": Counter(companies).most_common(5)
        }
        
    except Exception as e:
        return {"error": str(e)}


def get_applications_summary(file_path: str) -> str:
    """
    Get a formatted summary of applications.
    
    Args:
        file_path: Path to Excel file
    
    Returns:
        Formatted summary string
    """
    stats = get_application_stats(file_path)
    
    if stats.get("error"):
        return f"Error: {stats['error']}"
    
    if stats["total"] == 0:
        return "No applications tracked yet."
    
    lines = [
        "=" * 50,
        "JOB APPLICATION TRACKER SUMMARY",
        "=" * 50,
        "",
        f"📊 Total Applications: {stats['total']}",
        f"🏢 Unique Companies: {stats['unique_companies']}",
        "",
        "📈 By Status:"
    ]
    
    for status, count in sorted(stats["by_status"].items(), key=lambda x: -x[1]):
        pct = count / stats["total"] * 100
        lines.append(f"   {status}: {count} ({pct:.0f}%)")
    
    lines.append("")
    lines.append("📍 By Source:")
    for source, count in sorted(stats["by_source"].items(), key=lambda x: -x[1]):
        lines.append(f"   {source}: {count}")
    
    if stats.get("top_companies"):
        lines.append("")
        lines.append("🏆 Top Companies Applied:")
        for company, count in stats["top_companies"]:
            lines.append(f"   {company}: {count}")
    
    lines.append("")
    lines.append("=" * 50)
    
    return "\n".join(lines)


# ============================================================================
# EXPORTS
# ============================================================================

__all__ = [
    "build_excel_writer_graph",
    "create_excel_writer_subgraph",
    "save_application_to_excel",
    "quick_save_application",
    "update_application_status",
    "get_application_stats",
    "get_applications_summary"
]
