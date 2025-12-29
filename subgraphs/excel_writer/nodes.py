"""
Excel Writer Nodes

Saves job application data to Excel tracking spreadsheet.

Nodes:
1. prepare_record - Create ApplicationRecord from inputs
2. load_or_create_workbook - Load existing or create new Excel file
3. write_record - Write application record to spreadsheet
4. format_spreadsheet - Apply formatting and styling
5. save_workbook - Save and close workbook
"""

import os
from typing import Dict, Any, Optional
from pathlib import Path
from datetime import datetime

# Excel libraries
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl not available. Install with: pip install openpyxl")

from subgraphs.excel_writer.state import (
    ExcelWriterState,
    ApplicationStatus,
    ApplicationSource,
    ApplicationRecord,
    SpreadsheetConfig,
    SPREADSHEET_COLUMNS,
    generate_application_id,
    get_today
)


# ============================================================================
# NODE 1: PREPARE RECORD
# ============================================================================

def prepare_record(state: ExcelWriterState) -> Dict[str, Any]:
    """
    Create ApplicationRecord from inputs.
    """
    jd = state.structured_jd
    resume = state.resume_json
    existing_record = state.application_record
    
    print(f"  📝 Preparing application record...")
    
    # Start with existing record or create new
    if existing_record:
        record = existing_record
    else:
        record = ApplicationRecord()
    
    # Generate ID if not set
    if not record.application_id:
        record.application_id = generate_application_id()
    
    # Fill from JD
    if jd:
        record.company = record.company or jd.company_name or ""
        record.role = record.role or jd.role_title or ""
        record.location = record.location or jd.location or ""
        record.job_url = record.job_url or getattr(jd, 'application_url', None) or ""
        record.salary_range = record.salary_range or getattr(jd, 'salary_range', None) or ""
        record.work_type = record.work_type or getattr(jd, 'employment_type', None) or ""
        
        # Key requirements
        if jd.skills_required and not record.key_requirements:
            record.key_requirements = ", ".join(jd.skills_required[:5])
    
    # Fill from resume
    if resume and resume.header:
        # Could use resume version tracking here
        pass
    
    # Set scores
    if state.resume_score > 0:
        record.resume_score = state.resume_score
    if state.cover_letter_score > 0:
        record.cover_letter_score = state.cover_letter_score
    
    # Set dates
    if not record.date_found:
        record.date_found = get_today()
    record.last_updated = get_today()
    
    # Set default status
    if record.status == ApplicationStatus.NOT_APPLIED:
        record.status = ApplicationStatus.NOT_APPLIED
    
    print(f"  ✅ Record prepared: {record.company} - {record.role}")
    
    return {"application_record": record}


# ============================================================================
# NODE 2: LOAD OR CREATE WORKBOOK
# ============================================================================

def load_or_create_workbook(state: ExcelWriterState) -> Dict[str, Any]:
    """
    Load existing workbook or create new one with headers.
    """
    if not OPENPYXL_AVAILABLE:
        return {"error_message": "openpyxl not installed. Run: pip install openpyxl"}
    
    config = state.config
    file_path = config.file_path
    
    print(f"  📂 Loading/creating workbook: {file_path}")
    
    # Ensure output directory exists
    output_dir = Path(file_path).parent
    if output_dir and str(output_dir) != ".":
        output_dir.mkdir(parents=True, exist_ok=True)
    
    try:
        if os.path.exists(file_path):
            # Load existing workbook
            wb = load_workbook(file_path)
            if config.sheet_name in wb.sheetnames:
                ws = wb[config.sheet_name]
                print(f"  ✅ Loaded existing workbook with {ws.max_row - 1} records")
            else:
                # Create new sheet
                ws = wb.create_sheet(config.sheet_name)
                _write_headers(ws)
                print(f"  ✅ Created new sheet: {config.sheet_name}")
        else:
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = config.sheet_name
            _write_headers(ws)
            print(f"  ✅ Created new workbook")
        
        # Store workbook in state for next node
        return {
            "file_path": file_path,
            "_wb": wb,
            "_ws": ws
        }
        
    except Exception as e:
        print(f"  ❌ Failed to load/create workbook: {e}")
        return {"error_message": str(e)}


def _write_headers(ws):
    """Write header row to worksheet."""
    for col_idx, header in enumerate(SPREADSHEET_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")


# ============================================================================
# NODE 3: WRITE RECORD
# ============================================================================

def write_record(state: ExcelWriterState) -> Dict[str, Any]:
    """
    Write application record to spreadsheet.
    """
    record = state.application_record
    
    if not record:
        return {"error_message": "No application record to write"}
    
    if not OPENPYXL_AVAILABLE:
        return {"error_message": "openpyxl not installed"}
    
    # Load workbook
    file_path = state.file_path or state.config.file_path
    try:
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb[state.config.sheet_name] if state.config.sheet_name in wb.sheetnames else wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = state.config.sheet_name
            _write_headers(ws)
    except Exception as e:
        return {"error_message": f"Failed to load workbook: {e}"}
    
    print(f"  📝 Writing record to spreadsheet...")
    
    # Find next row
    next_row = ws.max_row + 1
    
    # Map record to columns
    row_data = [
        record.application_id,
        record.company,
        record.role,
        record.location,
        record.job_url,
        record.date_found,
        record.date_applied,
        record.last_updated,
        record.status.value if isinstance(record.status, ApplicationStatus) else record.status,
        record.source.value if isinstance(record.source, ApplicationSource) else record.source,
        record.referral_name,
        record.salary_range,
        record.work_type,
        record.resume_version,
        record.cover_letter_version,
        f"{record.resume_score:.1f}%" if record.resume_score > 0 else "",
        f"{record.cover_letter_score:.1f}%" if record.cover_letter_score > 0 else "",
        record.recruiter_name,
        record.recruiter_email,
        record.hiring_manager,
        record.next_followup,
        record.last_contact_date,
        record.interview_dates,
        record.offer_amount,
        record.response_time_days if record.response_time_days > 0 else "",
        record.notes,
        record.key_requirements,
        record.why_interested
    ]
    
    # Write row
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=next_row, column=col_idx, value=value)
    
    # Save immediately to preserve for next node
    wb.save(file_path)
    
    print(f"  ✅ Record written to row {next_row}")
    
    return {
        "row_number": next_row,
        "file_path": file_path
    }


# ============================================================================
# NODE 4: FORMAT SPREADSHEET
# ============================================================================

def format_spreadsheet(state: ExcelWriterState) -> Dict[str, Any]:
    """
    Apply formatting and styling to spreadsheet.
    """
    config = state.config
    file_path = state.file_path or config.file_path
    
    if not OPENPYXL_AVAILABLE:
        return {"error_message": "openpyxl not installed"}
    
    if not os.path.exists(file_path):
        return {"error_message": "Workbook file not found"}
    
    print(f"  🎨 Formatting spreadsheet...")
    
    try:
        wb = load_workbook(file_path)
        ws = wb[config.sheet_name] if config.sheet_name in wb.sheetnames else wb.active
        
        # Set column widths
        default_widths = {
            "A": 15,  # Application ID
            "B": 20,  # Company
            "C": 30,  # Role
            "D": 15,  # Location
            "E": 40,  # Job URL
            "F": 12,  # Date Found
            "G": 12,  # Date Applied
            "H": 12,  # Last Updated
            "I": 15,  # Status
            "J": 15,  # Source
            "K": 15,  # Referral Name
            "L": 15,  # Salary Range
            "M": 15,  # Work Type
            "N": 15,  # Resume Version
            "O": 18,  # Cover Letter Version
            "P": 12,  # Resume Score
            "Q": 15,  # CL Score
            "R": 20,  # Recruiter Name
            "S": 25,  # Recruiter Email
            "T": 20,  # Hiring Manager
            "U": 12,  # Next Follow-Up
            "V": 15,  # Last Contact
            "W": 15,  # Interview Dates
            "X": 15,  # Offer Amount
            "Y": 12,  # Response Time
            "Z": 40,  # Notes
            "AA": 30,  # Key Requirements
            "AB": 30   # Why Interested
        }
        
        for col, width in default_widths.items():
            ws.column_dimensions[col].width = width
        
        # Alternate row colors
        light_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        for row_idx in range(2, ws.max_row + 1):
            if row_idx % 2 == 0:
                for col_idx in range(1, len(SPREADSHEET_COLUMNS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = light_fill
        
        # Freeze header row
        if config.freeze_header:
            ws.freeze_panes = "A2"
        
        # Add auto-filter
        if config.enable_filter and ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(SPREADSHEET_COLUMNS))}{ws.max_row}"
        
        # Center align certain columns
        center_cols = [1, 6, 7, 8, 9, 10, 16, 17, 21, 22, 25]  # ID, dates, status, scores
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in center_cols:
                if col_idx <= len(SPREADSHEET_COLUMNS):
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center")
        
        # Save formatting
        wb.save(file_path)
        
        print(f"  ✅ Formatting applied")
        
    except Exception as e:
        print(f"  ⚠️ Formatting warning: {e}")
    
    return {"file_path": file_path}


# ============================================================================
# NODE 5: SAVE WORKBOOK
# ============================================================================

def save_workbook(state: ExcelWriterState) -> Dict[str, Any]:
    """
    Final save and verification.
    """
    file_path = state.file_path or state.config.file_path
    
    print(f"  💾 Verifying workbook: {file_path}")
    
    if not os.path.exists(file_path):
        return {"error_message": "Workbook file not found after write"}
    
    # Verify the file is valid
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        row_count = ws.max_row - 1  # Exclude header
        print(f"  ✅ Workbook verified: {row_count} total records")
        
        return {
            "file_path": file_path,
            "write_complete": True
        }
        
    except Exception as e:
        print(f"  ❌ Verification failed: {e}")
        return {"error_message": str(e)}


# ============================================================================
# EXPORTS
# ============================================================================

__all__ = [
    "prepare_record",
    "load_or_create_workbook",
    "write_record",
    "format_spreadsheet",
    "save_workbook"
]
